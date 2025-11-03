from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet 
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from typing import List, Dict, Any, Tuple, cast

class ExcelReportGenerator:
    @staticmethod
    def _setup_worksheet(wb: Workbook, title: str, color: str) -> Tuple[Worksheet, Font, PatternFill]:
        """Configura una hoja de trabajo con estilos básicos"""
        if wb.active is None:
            ws = wb.create_sheet()
        else:
            ws = cast(Worksheet, wb.active)
        
        ws.title = title
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        return ws, header_font, header_fill

    @staticmethod
    def _write_headers(ws: Worksheet, headers: List[str], font: Font, fill: PatternFill) -> None:
        """Escribe los encabezados en la hoja de trabajo con el estilo especificado"""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col)
            if not isinstance(cell, MergedCell):  # Solo modificar celdas normales
                cell.value = header
                cell.font = font
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center')

    @staticmethod
    def _adjust_column_widths(ws: Worksheet) -> None:
        """Ajusta el ancho de las columnas basado en el contenido"""
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    @staticmethod
    def generate_inventory_summary(data: List[Dict[str, Any]], periodo: str, sucursal_id: int) -> Workbook:
        wb = Workbook()
        ws, header_font, header_fill = ExcelReportGenerator._setup_worksheet(wb, "Resumen de Inventario", "366092")
        
        # Encabezados
        ws['A1'] = "Reporte: Resumen de Inventario"
        ws['A2'] = f"Período: {periodo}"
        ws['A3'] = f"Sucursal: {sucursal_id}"
        ws['A4'] = f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        headers = ['Código', 'Producto', 'Ubicación', 'Stock Actual', 'Stock Mínimo', 'Valor Total']
        ExcelReportGenerator._write_headers(ws, headers, header_font, header_fill)

        # Contenido
        for row, item in enumerate(data, 7):
            ws.cell(row=row, column=1, value=str(item.get('codigo_producto', '')))
            ws.cell(row=row, column=2, value=str(item.get('nombre_producto', '')))
            ws.cell(row=row, column=3, value=str(item.get('ubicacion', '')))
            ws.cell(row=row, column=4, value=item.get('stock_actual', 0))
            ws.cell(row=row, column=5, value=item.get('stock_minimo', 0))
            ws.cell(row=row, column=6, value=item.get('valor_total', 0))

        ExcelReportGenerator._adjust_column_widths(ws)
        return wb

    @staticmethod
    def generate_low_stock_report(data: List[Dict[str, Any]], periodo: str, sucursal_id: int) -> Workbook:
        wb = Workbook()
        ws, header_font, header_fill = ExcelReportGenerator._setup_worksheet(wb, "Stock Bajo", "C00000")
        
        # Encabezados
        ws['A1'] = "Reporte: Productos con Stock Bajo"
        ws['A2'] = f"Período: {periodo}"
        ws['A3'] = f"Sucursal: {sucursal_id}"
        ws['A4'] = f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        headers = ['Código', 'Producto', 'Ubicación', 'Stock Actual', 'Stock Mínimo', 'Diferencia']
        ExcelReportGenerator._write_headers(ws, headers, header_font, header_fill)

        # Contenido
        for row, item in enumerate(data, 7):
            ws.cell(row=row, column=1, value=str(item.get('codigo_producto', '')))
            ws.cell(row=row, column=2, value=str(item.get('nombre_producto', '')))
            ws.cell(row=row, column=3, value=str(item.get('ubicacion', '')))
            ws.cell(row=row, column=4, value=item.get('stock_actual', 0))
            ws.cell(row=row, column=5, value=item.get('stock_minimo', 0))
            ws.cell(row=row, column=6, value=item.get('diferencia', 0))

        ExcelReportGenerator._adjust_column_widths(ws)
        return wb

    @staticmethod
    def generate_movement_summary(data: List[Dict[str, Any]], periodo: str, sucursal_id: int) -> Workbook:
        wb = Workbook()
        ws, header_font, header_fill = ExcelReportGenerator._setup_worksheet(wb, "Mayores Movimientos", "548235")
        
        # Encabezados
        ws['A1'] = "Reporte: Resumen de Movimientos"
        ws['A2'] = f"Período: {periodo}"
        ws['A3'] = f"Sucursal: {sucursal_id}"
        ws['A4'] = f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        headers = ['Código', 'Producto', 'Total Entradas', 'Total Salidas', 'Movimientos Totales']
        ExcelReportGenerator._write_headers(ws, headers, header_font, header_fill)

        # Contenido
        for row, item in enumerate(data, 7):
            ws.cell(row=row, column=1, value=str(item.get('codigo_producto', '')))
            ws.cell(row=row, column=2, value=str(item.get('nombre_producto', '')))
            ws.cell(row=row, column=3, value=item.get('total_entradas', 0))
            ws.cell(row=row, column=4, value=item.get('total_salidas', 0))
            ws.cell(row=row, column=5, value=item.get('movimientos_totales', 0))

        ExcelReportGenerator._adjust_column_widths(ws)
        return wb